"""The scraped-data index: one row per source CSV, with coverage + provenance for
merging."""

import csv
import json

import pandas as pd
from openpyxl import load_workbook

from leaderspeech.text_scraper import index
from leaderspeech.text_scraper.run import SCHEMA_COLUMNS

RECIPE_YAML = r"""
source_id: arg_casarosada
country: Argentina
source_language: Spanish
start_urls: ["https://www.casarosada.gob.ar/discursos"]
listing: { link_selector: "a", link_pattern: "/discursos/\\d+" }
pagination: { type: query_param, param: page }
title: { selectors: ["h1"] }
text: { selectors: ["article"] }
date: { selectors: [".date"] }
"""


def _write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SCHEMA_COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in SCHEMA_COLUMNS})


def test_build_index_summarizes_a_source(tmp_path):
    recipes_dir = tmp_path / "recipes"
    recipes_dir.mkdir()
    (recipes_dir / "arg_casarosada.yml").write_text(RECIPE_YAML, encoding="utf-8")

    out_root = tmp_path / "scraped"
    _write_csv(out_root / "Argentina" / "arg_casarosada.csv", [
        {"doc_id": "ARG0001", "country": "Argentina", "date": "2020-01-01",
         "text": "uno", "source": "https://www.casarosada.gob.ar/discursos/1"},
        {"doc_id": "ARG0002", "country": "Argentina", "date": "2026-06-23",
         "text": "dos", "source": "https://www.casarosada.gob.ar/discursos/2"},
        {"doc_id": "ARG0003", "country": "Argentina", "date": "0001-11-30",  # bad year
         "text": "tres", "source": "https://www.casarosada.gob.ar/discursos/3"},
    ])
    # an _errors.csv sibling must be ignored
    (out_root / "Argentina" / "arg_casarosada_errors.csv").write_text(
        "timestamp,url,error\n", encoding="utf-8")

    # the harvested-link universe: a run's list, plus a probe snapshot that overlaps it,
    # carries one more link, and one URL the recipe's link_pattern no longer harvests
    (out_root / "Argentina" / "arg_casarosada_links.txt").write_text(
        "https://www.casarosada.gob.ar/discursos/1\n"
        "https://casarosada.gob.ar/discursos/2/\n"       # www + trailing slash -> dedupes
        "https://www.casarosada.gob.ar/discursos/3\n"
        "https://www.casarosada.gob.ar/discursos/4\n", encoding="utf-8")
    sample = out_root / "Argentina" / "sample"
    sample.mkdir(parents=True)
    (sample / "arg_casarosada_probe_20260101-000000.txt").write_text(
        "https://www.casarosada.gob.ar/discursos/4\n"    # overlaps -> union dedupes it
        "https://www.casarosada.gob.ar/discursos/5\n"
        "https://www.casarosada.gob.ar/galeria/7\n",     # stale-wide -> pruned by the pattern
        encoding="utf-8")
    (sample / "arg_casarosada_probe_20260101-000000.json").write_text(json.dumps(
        {"listing": {"mode": "spread (full history)", "links_found": 3,
                     "stopped_early": False, "stop_reason": "no_next_link"}}), encoding="utf-8")

    path = index.build_index(str(out_root), str(recipes_dir))
    assert path is not None

    df = pd.read_excel(path)
    assert list(df.columns) == index.COLUMNS
    assert len(df) == 1  # the _errors.csv was skipped
    row = df.iloc[0]
    assert row["source_id"] == "arg_casarosada"
    assert row["country"] == "Argentina"
    assert row["main_website"] == "www.casarosada.gob.ar"
    assert row["pagination_type"] == "query_param"
    assert row["n_speeches"] == 3
    assert row["date_min"] == "2020-01-01"          # the 0001 date is clipped out
    assert row["date_max"] == "2026-06-23"
    assert row["n_bad_or_missing_date"] == 1        # the 0001 date flagged
    assert row["doc_id_first"] == "ARG0001"
    assert row["doc_id_last"] == "ARG0003"
    assert row["iso3_prefix"] == "ARG"
    assert row["csv_file"].endswith("Argentina/arg_casarosada.csv")

    # the link columns, sited where the researcher asked for them
    assert index.COLUMNS.index("n_unique_links") == index.COLUMNS.index("n_speeches") + 1
    assert row["n_unique_links"] == 5        # 4 + 3, one overlapping, one /galeria/ pruned
    assert row["percent_scraped"] == 60.0    # 3 rows of 5 known
    assert row["links_status"] == "complete"
    assert row["links_last_harvested"]


def test_build_index_no_csvs_returns_none(tmp_path):
    assert index.build_index(str(tmp_path / "scraped"), str(tmp_path / "recipes")) is None


def test_index_blanks_the_link_columns_when_there_is_no_link_record(tmp_path):
    """arg_casarosada / chl_presidencia were scraped before link lists were saved. A `0`
    here would read as "we know of zero links" — the opposite of the truth."""
    recipes_dir = tmp_path / "recipes"
    recipes_dir.mkdir()
    (recipes_dir / "arg_casarosada.yml").write_text(RECIPE_YAML, encoding="utf-8")
    out_root = tmp_path / "scraped"
    _write_csv(out_root / "Argentina" / "arg_casarosada.csv", [
        {"doc_id": "ARG0001", "country": "Argentina", "date": "2020-01-01", "text": "uno"},
    ])

    row = pd.read_excel(index.build_index(str(out_root), str(recipes_dir))).iloc[0]
    assert pd.isna(row["n_unique_links"])
    assert pd.isna(row["percent_scraped"])
    assert row["links_status"] in ("", None) or pd.isna(row["links_status"])


def test_index_flags_stale_links_when_rows_exceed_known_links(tmp_path):
    """Above 100% is reported raw, not clamped: it is the one condition that PROVES the
    denominator is wrong, and hiding it would hide the bug."""
    recipes_dir = tmp_path / "recipes"
    recipes_dir.mkdir()
    (recipes_dir / "arg_casarosada.yml").write_text(RECIPE_YAML, encoding="utf-8")
    out_root = tmp_path / "scraped"
    _write_csv(out_root / "Argentina" / "arg_casarosada.csv", [
        {"doc_id": f"ARG000{i}", "country": "Argentina", "date": "2020-01-01", "text": "x"}
        for i in range(1, 4)
    ])
    (out_root / "Argentina" / "arg_casarosada_links.txt").write_text(
        "https://www.casarosada.gob.ar/discursos/1\n"
        "https://www.casarosada.gob.ar/discursos/2\n", encoding="utf-8")

    row = pd.read_excel(index.build_index(str(out_root), str(recipes_dir))).iloc[0]
    assert row["percent_scraped"] == 150.0
    assert row["links_status"] == "stale_links"


def test_index_writes_the_link_count_as_a_clean_integer(tmp_path):
    """Only openpyxl catches a float-typed cell rendering as "2.0" — pd.read_excel
    normalizes it away."""
    recipes_dir = tmp_path / "recipes"
    recipes_dir.mkdir()
    (recipes_dir / "arg_casarosada.yml").write_text(RECIPE_YAML, encoding="utf-8")
    out_root = tmp_path / "scraped"
    for country, sid in (("Argentina", "arg_casarosada"), ("Chile", "chl_presidencia")):
        _write_csv(out_root / country / f"{sid}.csv", [
            {"doc_id": "AAA0001", "country": country, "date": "2020-01-01", "text": "x"}])
    (out_root / "Argentina" / "arg_casarosada_links.txt").write_text(
        "https://www.casarosada.gob.ar/discursos/1\n"
        "https://www.casarosada.gob.ar/discursos/2\n", encoding="utf-8")

    path = index.build_index(str(out_root), str(recipes_dir))
    ws = load_workbook(path)["sources"]
    column = index.COLUMNS.index("n_unique_links") + 1
    cells = [ws.cell(row=r, column=column).value for r in range(2, ws.max_row + 1)]
    assert sorted(cells, key=lambda v: (v is None, v)) == [2, None]
    assert isinstance(cells[0] if cells[0] is not None else cells[1], int)


def test_index_lists_harvested_but_never_scraped_sources_on_a_second_sheet(tmp_path):
    """A source with links and no CSV is invisible to a `*/*.csv` glob. It gets its own
    sheet rather than a blank-`csv_file` row, which would trip merge.py's stale-index
    warning on every merge."""
    recipes_dir = tmp_path / "recipes"
    recipes_dir.mkdir()
    (recipes_dir / "arg_casarosada.yml").write_text(RECIPE_YAML, encoding="utf-8")
    out_root = tmp_path / "scraped"
    _write_csv(out_root / "Argentina" / "arg_casarosada.csv", [
        {"doc_id": "ARG0001", "country": "Argentina", "date": "2020-01-01", "text": "uno"}])
    (out_root / "Argentina" / "arg_casarosada_links.txt").write_text(
        "https://www.casarosada.gob.ar/discursos/1\n", encoding="utf-8")
    # harvested, never scraped — and a wayback_extend list that must NOT become a source
    (out_root / "Argentina" / "arg_otro_wayback_links.txt").write_text(
        "https://www.casarosada.gob.ar/discursos/8\n"
        "https://www.casarosada.gob.ar/discursos/9\n", encoding="utf-8")
    (out_root / "Argentina" / "arg_casarosada_wayback_extend_links.txt").write_text(
        "https://www.casarosada.gob.ar/discursos/7\n", encoding="utf-8")

    path = index.build_index(str(out_root), str(recipes_dir))
    assert pd.read_excel(path).shape[0] == 1          # sheet 0 is unchanged
    pending = pd.read_excel(path, sheet_name=index.SHEET_PENDING)
    assert list(pending.columns) == index.PENDING_COLUMNS
    assert list(pending["source_id"]) == ["arg_otro_wayback"]
    assert pending.iloc[0]["n_unique_links"] == 2
    assert pd.isna(pending.iloc[0]["recipe_file"])    # no recipe on disk for it
